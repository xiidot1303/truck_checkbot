from app.views import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.services.task_service import *

async def task_report_by_date(request):
    start_date = request.GET.get('created_at__range__gte')
    end_date = request.GET.get('created_at__range__lte')
    start_date = "10.10.2020" if not start_date or start_date == 'None' else start_date
    end_date = "10.10.2100" if not end_date or end_date == 'None' else end_date

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers (based on your data)
    headers = [
        ["Period", "Yo'nalish", "Mashina", "Haydovchi", "Mashina skladda bo'lishi", "", "", "Yuk ortilishi tugatilishi", "", "", "Manzilga yetib borish jarayoni", "", "", "Filialning yukni tushirish jarayoni", "", "", "Mashinaning zavodga qaytib kelishi", "", "", "Grafikdan jami kechikish"],
        ["", "", "", "", "Grafik", "Fact", "Farq", "Grafik", "Fact", "Farq",
        "Grafik", "Fact", "Farq", "Grafik", "Fact", "Farq", "Grafik", "Fact", "Farq", ""]
        ]

    # Add the headers to the worksheet
    for header in headers:
        ws.append(header)
    
    # Query the database
    tasks = await filter_completed_tasks_by_date_range(start_date, end_date)  # Replace 'YourModel' with your actual model name

    # Example: Add your tasks to the worksheet
    last_period = date(1900, 1, 1)
    async for task in tasks:
        task: Task
        async for taskdepot in task.depots.filter(): 
            taskdepot: TaskDepot
            row_data = [
                task.created_at.strftime("%d.%m.%Y") if last_period != task.created_at.date() else "", 
                taskdepot.branch, 
                await task.get_car, await task.get_driver
            ]
            total_differences = 0
            async for event in task.events.filter().order_by('id'):
                event: TaskEvent
                difference = await event.difference_with_schedule
                if difference < 0:
                    total_differences += difference
                data = [
                    event.schedule_time.strftime("%H:%M") if event.schedule_time else "",
                    event.end_time.strftime('%H:%M') if event.end_time else "",
                    difference
                ]
                row_data += data
            # add empty values in fields "Mashinaning zavodga qaytib kelishi"
            row_data += ["", "", ""]
            row_data += [str(total_differences)]
            ws.append(row_data)

        last_period = task.created_at.date()
    # Style the headers
    header_font = Font(bold=True)
    for col in range(1, 21):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths for better readability
    column_widths = [12, 12, 12, 20, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 12, 12, 12, 25]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Merge cells
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
    ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
    ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=16)
    ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=19)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
    ws.merge_cells(start_row=1, start_column=20, end_row=2, end_column=20)

    # Define border style
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Apply borders to the entire table
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_style

    # color the fields
    for row in range(1, ws.max_row + 1):
        for col in range(5, 8):  # Columns E, F, G
            ws.cell(row=row, column=col).fill = PatternFill(start_color="e2f0d9", end_color="e2f0d9", fill_type="solid")
        for col in range(8, 11):  # Columns G, H, J
            ws.cell(row=row, column=col).fill = PatternFill(start_color="dae3f4", end_color="dae3f4", fill_type="solid")
        for col in range(11, 14):  # Columns K, L, M
            ws.cell(row=row, column=col).fill = PatternFill(start_color="ffdae3", end_color="ffdae3", fill_type="solid")
        for col in range(14, 17):  # Columns N, O, P
            ws.cell(row=row, column=col).fill = PatternFill(start_color="f4ffda", end_color="f4ffda", fill_type="solid")
        for col in range(17, 20):  # Columns Q, R, S
            ws.cell(row=row, column=col).fill = PatternFill(start_color="e5d6ff", end_color="e5d6ff", fill_type="solid")
        for col in range(20, 21):  # Columns T
            ws.cell(row=row, column=col).fill = PatternFill(start_color="fffbe5", end_color="fffbe5", fill_type="solid")

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="schedule_report.xlsx"'
    wb.save(response)
    return response